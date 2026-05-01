from datetime import date
import random
import os
import csv
import re
from django.http import HttpResponse
from .models import Cliente

from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login as auth_login, logout as auth_logout
from django.contrib.auth.models import Group
from django.db import transaction
from django.db.utils import ProgrammingError
from django.db.models import Count, ProtectedError
from django.db.models.functions import TruncMonth
from django.shortcuts import redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.shortcuts import get_object_or_404
from django.http import HttpResponseForbidden
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth import get_user_model


from .models import (
    Cliente,
    Deposito,
    PerfilUsuario,
    SolicitudCorreccion,
    SolicitudCorreccionCliente,
    Vehiculo,
)


ROLE_ADMIN_MASTER = "admin_master"
ROLE_ADMIN = "administrador"
ROLE_OPERADOR = "operador"
ROLE_PROMOTOR = "promotor"
ALLOWED_EMAIL_DOMAINS = [
    item.strip().lower()
    for item in (os.getenv("ALLOWED_EMAIL_DOMAINS", "") or "").split(",")
    if item.strip()
]

ROLE_LABELS = {
    ROLE_ADMIN_MASTER: "Admin Master",
    ROLE_ADMIN: "Administrador",
    ROLE_OPERADOR: "Operador",
    ROLE_PROMOTOR: "Promotor",
}

_ADMIN_PERMS = {
    "ver_dashboard",
    "ver_inventario",
    "operadorregistrador",
    "liberar",
    "gestionar_depositos",
    "gestionar_correcciones",
    "gestionar_correcciones_clientes",
    "gestionar_usuarios",
    "solicitar_correccion",
    "solicitar_correccion_cliente",
    "gestionar_credito",
    "editar_cliente"
}

ROLE_PERMISSIONS = {
    ROLE_ADMIN_MASTER: _ADMIN_PERMS | {"auditar_admin", "buscar_por_id"},
    ROLE_ADMIN: set(_ADMIN_PERMS),
    ROLE_OPERADOR: {
        "ver_dashboard",
        "operadorregistrador",
        "ver_inventario",
        "solicitar_correccion",
        "solicitar_correccion_cliente",
        "editar_cliente",
    },
    ROLE_PROMOTOR: {"ver_dashboard", 
                    "ver_inventario",
                    "operadorregistrador",
                    },
}

CORRECCION_FIELDS = {
    "fecha_ingreso": "Fecha de ingreso",
    "deposito": "Deposito",
    "oficio": "Numero de oficio",
    "fecha_oficio": "Fecha de oficio",
    "titular": "Titular",
    "observaciones": "Observaciones",
}

CLIENTE_CORRECCION_FIELDS = {
    "sap": "Codigo SAP",
    "nombre": "Nombre",
    "tipo_cuenta": "Tipo de cuenta",
    "lista_precios": "Lista de precios",
    "latitud": "Latitud",
    "longitud": "Longitud",
    "direccion": "Direccion",
    "zona": "Zona",
    "estado": "Estado",
    "poblacion": "Poblacion",
}

EDITABLE_FIELDS_OPERADOR = [
    'lista_precios',
    'tipo_cuenta',
    'poblacion',
    'estado',
    'zona'
]


def _normalize_excel_header(value):
    import unicodedata

    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return (
        text.replace("\u00a0", " ")
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
    )

def editar_cliente(request, cliente_id):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "editar_cliente"):
        return _reject_unauthorized(request)

    cliente = get_object_or_404(Cliente, id=cliente_id)
    role = _get_role(request)
    user = _get_current_user(request)

    # Seguridad: operador solo edita sus clientes
    if role == ROLE_OPERADOR and cliente.operador != user:
        return HttpResponseForbidden("No tienes permiso para editar este cliente.")

    if request.method == 'POST':
        if role == ROLE_OPERADOR:
            if cliente.edicion_operador_usada:
                messages.error(request, "Ya usaste tu única edición permitida para este cliente.")
                return redirect('clientes_list')

            for field in EDITABLE_FIELDS_OPERADOR:
                if field in request.POST:
                    setattr(cliente, field, (request.POST.get(field) or "").strip())

            cliente.edicion_operador_usada = True
            cliente.save(update_fields=EDITABLE_FIELDS_OPERADOR + ['edicion_operador_usada', 'actualizado_en'])
            messages.success(request, f"Cliente {cliente.sap} actualizado correctamente (edición única usada).")
            return redirect('clientes_list')

        if role in [ROLE_ADMIN, ROLE_ADMIN_MASTER]:
            # Admin / Admin Master pueden editar libremente solo campos de negocio permitidos aquí
            admin_editable_fields = [
                'lista_precios', 'tipo_cuenta', 'poblacion', 'estado', 'zona',
                'nombre', 'direccion', 'latitud', 'longitud', 'frecuencia_visita', 'dias_visita'
            ]
            touched = []
            for field in admin_editable_fields:
                if field in request.POST:
                    if field == 'dias_visita':
                        setattr(cliente, field, ','.join(request.POST.getlist(field)))
                    else:
                        setattr(cliente, field, (request.POST.get(field) or "").strip())
                    touched.append(field)

            if touched:
                cliente.save(update_fields=touched + ['actualizado_en'])
                messages.success(request, f"Cliente {cliente.sap} actualizado correctamente.")
            else:
                messages.info(request, "No se detectaron cambios para guardar.")
            return redirect('clientes_list')

    return render(request, 'Vehiculos/editar_cliente.html', {
        'cliente': cliente,
        'rol': role,
        'editable_fields_operador': EDITABLE_FIELDS_OPERADOR,
        'edicion_operador_usada': cliente.edicion_operador_usada,
    })


def _canonicalize_import_header(normalized_header):
    header = normalized_header or ""

    if header.startswith("codigo_sa"):
        return "sap"
    if header in ("sap", "codigo_sap"):
        return "sap"

    if header.startswith("tipo_cuen"):
        return "tipo_cuenta"
    if header in ("tipo_cuenta", "tipo_de_cuenta"):
        return "tipo_cuenta"

    if header in ("nombre",):
        return "nombre"

    if header in ("direccion",):
        return "direccion"

    if header in ("zona",):
        return "zona"

    if header in ("estado",):
        return "estado"

    if header in ("poblacion",):
        return "poblacion"

    if header in ("latitud",):
        return "latitud"

    if header in ("longitud",):
        return "longitud"

    if header.startswith("lista_de_pre"):
        return "lista_precios"
    if header in ("lista_precios", "lista_de_precios"):
        return "lista_precios"

    if header in ("fecha_registro", "fecha_de_registro", "fecha_ingreso", "fecha_de_ingreso"):
        return "fecha_registro"

    return header


def _parse_excel_date(value):
    if value is None or value == "":
        return None

    from datetime import datetime

    if isinstance(value, (date,)) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    raw = str(value).strip()

    parsed = parse_date(raw)
    if parsed:
        return parsed

    # Soportar DD/MM/YYYY
    if "/" in raw:
        parts = raw.split("/")
        if len(parts) == 3:
            try:
                day = int(parts[0])
                month = int(parts[1])
                year = int(parts[2])
                return date(year, month, day)
            except Exception:
                return None
    return None


def _is_logged_in(request):
    return "usuario" in request.session


def _get_role(request):
    role = request.session.get("rol")
    return role if role in ROLE_LABELS else ROLE_PROMOTOR


def _get_role_for_user(user):
    if user.groups.filter(name=ROLE_ADMIN_MASTER).exists():
        return ROLE_ADMIN_MASTER
    if user.is_superuser:
        return ROLE_ADMIN
    if user.groups.filter(name=ROLE_ADMIN).exists():
        return ROLE_ADMIN
    if user.groups.filter(name=ROLE_OPERADOR).exists():
        return ROLE_OPERADOR
    if user.groups.filter(name=ROLE_PROMOTOR).exists():
        return ROLE_PROMOTOR
    return ROLE_PROMOTOR


def _get_current_user(request):
    username = request.session.get("usuario")
    if not username:
        return None
    User = get_user_model()
    return User.objects.filter(username=username).first()


def _normalize_email(value):
    return (value or "").strip().lower()


def _email_allowed(email):
    normalized = _normalize_email(email)
    if ALLOWED_EMAIL_DOMAINS and not any(normalized.endswith(domain) for domain in ALLOWED_EMAIL_DOMAINS):
        return False
    if normalized.count("@") != 1:
        return False
    local_part = normalized.split("@", 1)[0]
    return bool(local_part)


def _get_role_groups():
    groups = {}
    for role in ROLE_LABELS:
        group, _ = Group.objects.get_or_create(name=role)
        groups[role] = group
    return groups


def _role_to_prefijo(role):
    if role == ROLE_ADMIN_MASTER:
        return PerfilUsuario.PREFIJO_ADMIN_MASTER
    if role == ROLE_ADMIN:
        return PerfilUsuario.PREFIJO_ADMINISTRADOR
    if role == ROLE_OPERADOR:
        return PerfilUsuario.PREFIJO_OPERADOR
    if role == ROLE_PROMOTOR:
        return PerfilUsuario.PREFIJO_PROMOTOR

    return PerfilUsuario.PREFIJO_PROMOTOR




def _generar_numero_empleado(role):
    prefijo = _role_to_prefijo(role)
    last = (
        PerfilUsuario.objects.filter(numero_interno__startswith=f"{prefijo}-")
        .order_by("-numero_interno")
        .first()
    )
    last_num = 0
    if last and "-" in (last.numero_interno or ""):
        try:
            last_num = int(last.numero_interno.split("-", 1)[1])
        except ValueError:
            last_num = 0

    for candidate_num in range(last_num + 1, last_num + 100000):
        candidate = f"{prefijo}-{candidate_num:05d}"
        if not PerfilUsuario.objects.filter(numero_interno=candidate).exists():
            return candidate
    raise ValueError("No se pudo generar un numero de empleado unico.")

def generar_codigo_sap():
    last = (
        Cliente.objects
        .filter(codigo_sap__startswith="SAP-")
        .order_by("-codigo_sap")
        .first()
    )

    last_num = 0
    if last and last.codigo_sap:
        try:
            last_num = int(last.codigo_sap.split("-")[1])
        except:
            last_num = 0

    for num in range(last_num + 1, last_num + 100000):
        codigo = f"SAP-{num:05d}"
        if not Cliente.objects.filter(codigo_sap=codigo).exists():
            return codigo

    raise ValueError("No se pudo generar un código SAP único")

def _validate_pdf(file_obj, label):
    if not file_obj:
        return f"Falta el PDF de {label}."
    name = (getattr(file_obj, "name", "") or "").lower()
    content_type = (getattr(file_obj, "content_type", "") or "").lower()
    if not name.endswith(".pdf"):
        return f"El archivo de {label} debe ser PDF."
    if content_type and content_type != "application/pdf":
        return f"El archivo de {label} debe ser PDF."
    return None


def _coerce_model_field_value(model_cls, field_name, raw_value):
    field = model_cls._meta.get_field(field_name)
    if raw_value is None:
        return None
    value = raw_value.strip()

    internal = field.get_internal_type()
    if internal in ("IntegerField", "PositiveIntegerField", "BigIntegerField", "SmallIntegerField"):
        return int(value)
    if internal == "DateField":
        parsed = parse_date(value)
        if not parsed:
            raise ValueError("Fecha invalida. Usa formato YYYY-MM-DD.")
        return parsed
    if internal in ("DecimalField", "FloatField"):
        return float(value)
    if internal == "BooleanField":
        return value.lower() in ("1", "true", "si", "sí", "on")
    return value


def _coerce_field_value(field_name, raw_value):
    field = Vehiculo._meta.get_field(field_name)
    if raw_value is None:
        return None
    value = raw_value.strip()

    if field.get_internal_type() in ("IntegerField", "PositiveIntegerField", "BigIntegerField", "SmallIntegerField"):
        return int(value)
    if field.get_internal_type() == "DateField":
        parsed = parse_date(value)
        if not parsed:
            raise ValueError("Fecha invalida. Usa formato YYYY-MM-DD.")
        return parsed
    return value


def _generate_folio():
    today = timezone.localdate()
    prefix = today.strftime("Number-%y%m%d-")
    for _ in range(25):
        serial = random.randint(100, 999)
        folio = f"{prefix}{serial}"
        if not Vehiculo.objects.filter(folio=folio).exists():
            return folio
    return f"{prefix}{random.randint(1000, 9999)}"


def _get_folio_sugerido(request):
    folio = request.session.get("folio_sugerido")
    if folio and not Vehiculo.objects.filter(folio=folio).exists():
        return folio
    folio = _generate_folio()
    request.session["folio_sugerido"] = folio
    return folio

def _has_permission(request, permission):
    return permission in ROLE_PERMISSIONS.get(_get_role(request), set())


def _scoped_clientes_queryset(request):
    role = _get_role(request)
    user = _get_current_user(request)

    if role == ROLE_OPERADOR:
        if user is None:
            return Cliente.objects.none()
        return Cliente.objects.filter(
            Q(operador=user) |
            Q(operador__perfil__operador_asignado=user)
        )

    if role == ROLE_PROMOTOR:
        if user is None:
            return Cliente.objects.none()
        return Cliente.objects.filter(operador=user)

    return Cliente.objects.all()

def _reject_unauthorized(request):
    messages.error(request, "Tu rol no tiene permiso para esta accion.")
    return redirect("dashboard")

def _generar_codigo_sap():
    # Usamos values_list para extraer SOLO el texto del SAP.
    # Así evitamos leer toda la fila y evitamos que SQLite explote con los decimales corruptos antiguos.
    last_sap = Cliente.objects.filter(sap__startswith="SAP-").order_by("-sap").values_list("sap", flat=True).first()
    
    last_num = 0
    
    if last_sap and "-" in last_sap:
        try:
            # Extraemos la parte numérica (ej. de "SAP-00015" sacamos "15")
            last_num = int(last_sap.split("-", 1)[1])
        except ValueError:
            last_num = 0

    # Buscamos el siguiente número libre (sumando 1)
    for candidate_num in range(last_num + 1, last_num + 100000):
        candidate = f"SAP-{candidate_num:05d}"
        if not Cliente.objects.filter(sap=candidate).exists():
            return candidate
            
    raise ValueError("No se pudo generar un código SAP único.")

def login_view(request):
    error = None

    if request.method == 'POST':
        usuario = (request.POST.get('usuario') or '').strip()
        password = request.POST.get('password') or ''

        if usuario and password:
            if not usuario.endswith('@usuario.com'):
                usuario += '@usuario.com'
            user = authenticate(request, username=usuario, password=password)
            if user:
                auth_login(request, user)
                request.session['usuario'] = user.get_username()
                request.session['rol'] = _get_role_for_user(user)
                return redirect('dashboard')
            error = 'Credenciales invalidas. Verifica tu usuario y contraseña.'
        else:
            error = 'Ingresa usuario y contraseña para continuar.'

    return render(request, 'Vehiculos/login.html', {'error': error})


def dashboard(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "ver_dashboard"):
        return redirect('login')

    clientes_qs = _scoped_clientes_queryset(request)

    total = clientes_qs.count()

    directos = clientes_qs.filter(tipo_cuenta=Cliente.TIPO_DIRECTO).count()
    prospectos = clientes_qs.filter(tipo_cuenta=Cliente.TIPO_PROSPECTO).count()

    pendientes = prospectos
    liberados = directos
    en_proceso = 0

    monthly_data = list(
        reversed(
            clientes_qs.filter(fecha_registro__isnull=False)
            .annotate(month=TruncMonth('fecha_registro'))
            .values('month')
            .annotate(total=Count('id'))
            .order_by('-month')[:6]
        )
    )
    monthly_data = list(reversed(monthly_data))
    monthly_labels = [item['month'].strftime('%b') for item in monthly_data]
    monthly_ingress = [item['total'] for item in monthly_data]

    tipo_data = clientes_qs.values('tipo_cuenta').annotate(total=Count('id'))
    
    type_labels = [item['tipo_cuenta'] for item in tipo_data]
    type_values = [item['total'] for item in tipo_data]

    actividad = clientes_qs.order_by('-fecha_registro', '-id')[:5]

    context = {
    'resumen_data': {
    'total': total,
    'pendientes': pendientes,
    'liberados': liberados,
    'enProceso': en_proceso,
    },
    'detalle_data': {
        'listas': list(clientes_qs.values('lista_precios').annotate(total=Count('id'))),
        'labels': [item['lista_precios'] for item in clientes_qs.values('lista_precios').annotate(total=Count('id'))],
        'values': [item['total'] for item in clientes_qs.values('lista_precios').annotate(total=Count('id'))],
        'monthlyLabels': monthly_labels,
        'monthlyIngress': monthly_ingress,
        'typeLabels': type_labels,
        'typeValues': type_values,
    },
    'actividad': actividad,
    'hoy': date.today(),
    'rol': _get_role(request),
    'rol_label': ROLE_LABELS[_get_role(request)],
    'can_operador': _has_permission(request, "operadorregistrador"),
    'can_clientes': _has_permission(request, "operadorregistrador") or _has_permission(request, "gestionar_usuarios"),
    'can_historial': _has_permission(request, "operadorregistrador"),
    'can_depositos': _has_permission(request, "gestionar_depositos"),
    'can_correcciones': _has_permission(request, "gestionar_correcciones_clientes"),
    'can_usuarios': _has_permission(request, "gestionar_usuarios"),
    }
    return render(request, 'Vehiculos/dashboard.html', context)


def depositos_view(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "gestionar_depositos"):
        return _reject_unauthorized(request)

    if request.method == 'POST':
        nombre = (request.POST.get('nombre') or '').strip()
        if not nombre:
            messages.error(request, 'Ingresa un nombre de deposito valido.')
            return redirect('depositos')
        if Deposito.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, f'El deposito "{nombre}" ya existe.')
            return redirect('depositos')

        Deposito.objects.create(nombre=nombre)
        messages.success(request, f'Deposito "{nombre}" agregado correctamente.')
        return redirect('depositos')

    depositos = list(Deposito.objects.order_by('nombre'))
    conteos = {
        item['deposito']: item['total']
        for item in Vehiculo.objects.values('deposito').annotate(total=Count('id'))
    }
    depositos_data = [
        {
            'nombre': deposito.nombre,
            'creado_en': deposito.creado_en,
            'total': conteos.get(deposito.nombre, 0),
        }
        for deposito in depositos
    ]
    return render(
        request,
        'Vehiculos/depositos.html',
        {
            'depositos': depositos_data,
            'can_depositos': True,
        },
    )


def usuarios_view(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "gestionar_usuarios"):
        return _reject_unauthorized(request)

    User = get_user_model()
    role_groups = _get_role_groups()
    role_names = list(ROLE_LABELS.keys())
    current_user = _get_current_user(request)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()

        if action == 'create':
            nombre_usuario = (request.POST.get('nombre_usuario') or '').strip()
            email = nombre_usuario + '@usuario.com'
            email = _normalize_email(email)
            password = request.POST.get('password') or ''
            telefono = (request.POST.get('telefono') or '').strip()
            nombre_completo = (request.POST.get('nombre_completo') or '').strip()
            role = (request.POST.get('role') or '').strip()
            rfc_pdf = request.FILES.get('rfc_pdf')
            ine_pdf = request.FILES.get('ine_pdf')
            comprobante_pdf = request.FILES.get('comprobante_domicilio_pdf')

            if not email or not password or not telefono or not nombre_completo or role not in ROLE_LABELS:
                messages.error(request, 'Completa nombre, correo, teléfono, contraseña y rol para crear la cuenta.')
                return redirect('usuarios')

            if not _email_allowed(email):
                messages.error(request, 'Solo se permiten correos con dominio @usuario.com.')
                return redirect('usuarios')

            if User.objects.filter(username=email).exists():
                messages.error(request, 'Ya existe una cuenta registrada con ese correo.')
                return redirect('usuarios')

            pdf_errors = [
                _validate_pdf(rfc_pdf, "RFC"),
                _validate_pdf(ine_pdf, "INE"),
                _validate_pdf(comprobante_pdf, "Comprobante de domicilio"),
            ]
            pdf_errors = [err for err in pdf_errors if err]
            if pdf_errors:
                for err in pdf_errors:
                    messages.error(request, err)
                return redirect('usuarios')

            with transaction.atomic():
                user = User.objects.create_user(username=email, email=email, password=password)
                user.first_name = nombre_completo
                role_group = role_groups[role]
                user.groups.remove(*Group.objects.filter(name__in=role_names))
                user.groups.add(role_group)
                if role == ROLE_ADMIN:
                    user.is_staff = True
                    user.is_superuser = True
                else:
                    user.is_staff = False
                    user.is_superuser = False
                user.save()

                operador_asignado = None

                if role == ROLE_PROMOTOR:
                    operador_id = request.POST.get('operador_asignado')

                    if not operador_id:
                        messages.error(request, "Selecciona un operador.")
                        return redirect('usuarios')

                    operador_asignado = User.objects.filter(
                        id=operador_id,
                        groups__name=ROLE_OPERADOR
                    ).first()

                    if operador_asignado is None:
                        messages.error(request, "El operador seleccionado no es válido.")
                        return redirect('usuarios')

                PerfilUsuario.objects.create(
                    user=user,
                    numero_interno=_generar_numero_empleado(role),
                    nombre_completo=nombre_completo,
                    telefono=telefono,
                    rfc_pdf=rfc_pdf,
                    ine_pdf=ine_pdf,
                    comprobante_domicilio_pdf=comprobante_pdf,
                    operador_asignado=operador_asignado,
                    contrasena_temporal=password,
                )
            messages.success(request, f'Cuenta {email} creada correctamente.')
            return redirect('usuarios')

        if action == 'delete':
            user_id = request.POST.get('user_id')
            target = User.objects.filter(id=user_id).first()
            if not target:
                messages.error(request, 'No se encontro la cuenta solicitada.')
                return redirect('usuarios')

            if current_user and target.id == current_user.id:
                messages.error(request, 'No puedes eliminar tu propia cuenta.')
                return redirect('usuarios')

            target_email = _normalize_email(target.email or target.username)
            if not _email_allowed(target_email):
                messages.error(request, 'Solo puedes eliminar cuentas con dominio @usuario.com.')
                return redirect('usuarios')

            try:
                target.delete()
                messages.success(request, f'Cuenta {target_email} eliminada correctamente.')
            except ProtectedError:
                clientes_count = Cliente.objects.filter(operador=target).count()
                messages.error(
                    request,
                    f'No se puede eliminar la cuenta {target_email}. Tiene {clientes_count} cliente(s) asociado(s). '
                    f'Por favor, reasigna los clientes a otro operador antes de eliminar esta cuenta.'
                )
            return redirect('usuarios')

        messages.error(request, 'Accion invalida.')
        return redirect('usuarios')

    usuarios = []
    all_users = list(User.objects.order_by('username'))
    try:
        perfiles = {
            perfil.user_id: perfil
            for perfil in PerfilUsuario.objects.filter(user__in=all_users)
        }
    except ProgrammingError:
        messages.error(request, "Faltan migraciones de PerfilUsuario. Ejecuta: manage.py migrate")
        perfiles = {}
    for user in all_users:
        display_email = (user.email or user.username or '').strip()
        if not _email_allowed(display_email):
            continue
        role = _get_role_for_user(user)
        perfil = perfiles.get(user.id)
        usuarios.append(
            {
                'id': user.id,
                'email': display_email,
                'nombre_completo': (perfil.nombre_completo if perfil else (user.first_name or '')).strip(),
                'numero_empleado': perfil.numero_interno if perfil else '',
                'telefono': perfil.telefono if perfil else '',
                'password': perfil.contrasena_temporal if perfil else '',
                'docs_ok': bool(
                    perfil
                    and perfil.rfc_pdf
                    and perfil.ine_pdf
                    and perfil.comprobante_domicilio_pdf
                ),
                'role': role,
                'role_label': ROLE_LABELS.get(role, role),
                'is_self': current_user and user.id == current_user.id,
            }
        )


    operadores = User.objects.filter(groups__name=ROLE_OPERADOR)

    return render(
        request,
        'Vehiculos/usuarios.html',
        {
            'usuarios': usuarios,
            'role_options': [(key, ROLE_LABELS[key]) for key in ROLE_LABELS],
            'operadores' : operadores
        },
    )


def registrar_vehiculo(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "registrar"):
        return _reject_unauthorized(request)

    def build_context():
        return {
            'depositos': list(Deposito.objects.order_by('nombre').values_list('nombre', flat=True)),
            'can_depositos': _has_permission(request, "gestionar_depositos"),
            'folio_sugerido': _get_folio_sugerido(request),
        }

    if request.method == 'POST':
        post = request.POST
        archivo = request.FILES.get('documento_pdf')

        required = [
            'fecha_ingreso',
            'autoridad',
            'deposito',
            'marca',
            'modelo',
            'anio',
            'placas',
            'vin',
            'numero_motor',
            'tipo_servicio',
            'estatus_legal',
            'grua_motivo',
            'grua_direccion',
        ]
        missing = [field for field in required if not (post.get(field) or '').strip()]
        if missing:
            missing_labels = [CORRECCION_FIELDS.get(field, field) for field in missing]
            messages.error(
                request,
                f'Completa los campos obligatorios: {", ".join(missing_labels)}.',
            )
            return render(request, 'Vehiculos/registrar-vehiculo.html', build_context())

        field_limits = {
            'turno': 20,
            'autoridad': 120,
            'deposito': 120,
            'motivo': 280,
            'grua_motivo': 240,
            'grua_direccion': 180,
            'marca': 60,
            'modelo': 60,
            'color': 40,
            'placas': 15,
            'vin': 12,
            'numero_motor': 40,
            'tipo_servicio': 30,
            'combustible': 20,
            'oficio': 80,
            'titular': 120,
            'observaciones': 280,
        }
        for field, max_len in field_limits.items():
            value = (post.get(field) or '').strip()
            if value and len(value) > max_len:
                label = CORRECCION_FIELDS.get(field, field)
                messages.error(request, f'El campo {label} excede {max_len} caracteres.')
                return render(request, 'Vehiculos/clientes.html', build_context())

        # vin_value = (post.get('vin') or '').strip()
        # if vin_value and len(vin_value) != 17:
        #     messages.error(request, 'El VIN debe tener exactamente 17 caracteres.')
        #     return render(request, 'Vehiculos/registrar-vehiculo.html', build_context())

        #  NumerodeCliente = request.session.get("Numero_de_Cliente_sugerido") or _generate_NumerodeCliente()
        #  NumerodeCliente = NumerodeCliente.strip().upper()
        #  if Vehiculo.objects.filter(NumerodeCliente=NumerodeCliente).exists():
        #      NumerodeCliente = _generate_NumerodeCliente().strip().upper()

        # try:
        #     anio = int(post.get('anio', '0'))
        #     kilometraje = int(post.get('kilometraje', '0') or 0)
        # except ValueError:
        #     messages.error(request, 'Revisa los campos numericos.')
        #     return render(request, 'Vehiculos/registrar-vehiculo.html', build_context())

        vehiculo = Vehiculo.objects.create(
            # NumerodeCliente=NumerodeCliente,
            fecha_ingreso=post.get('fecha_ingreso'),
            turno=(post.get('turno') or '').strip(),
            autoridad=(post.get('autoridad') or '').strip(),
            deposito=(post.get('deposito') or '').strip(),
            motivo=(post.get('motivo') or '').strip(),
            grua_motivo=(post.get('grua_motivo') or '').strip(),
            grua_direccion=(post.get('grua_direccion') or '').strip(),
            marca=(post.get('marca') or '').strip(),
            modelo=(post.get('modelo') or '').strip(),
            # anio=anio,
            # color=(post.get('color') or '').strip(),
            # placas=(post.get('placas') or '').strip().upper(),
            # vin=(post.get('vin') or '').strip().upper(),
            # numero_motor=(post.get('numero_motor') or '').strip().upper(),
            # tipo_servicio=(post.get('tipo_servicio') or '').strip(),
            # combustible=(post.get('combustible') or '').strip(),
            # kilometraje=kilometraje,
            estatus_legal=(post.get('estatus_legal') or Vehiculo.ESTATUS_EN_CUSTODIA).strip(),
            oficio=(post.get('oficio') or '').strip(),
            fecha_oficio=post.get('fecha_oficio') or None,
            titular=(post.get('titular') or '').strip(),
            observaciones=(post.get('observaciones') or '').strip(),
            documento_nombre=archivo.name if archivo else '',
            liberado=(post.get('estatus_legal') or '').strip() == Vehiculo.ESTATUS_LIBERADO,
        )

        if vehiculo.liberado:
            vehiculo.fecha_liberacion = timezone.now()
            vehiculo.save(update_fields=['fecha_liberacion'])

        request.session.pop("Numero_de_Cliente_sugerido", None)
        messages.success(request, f'Vehiculo {vehiculo.NumerodeCliente} registrado correctamente.')
        return redirect('vehiculos')

    return render(request, 'Vehiculos/registrar-vehiculo.html', build_context())


def vehiculos_list(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "ver_inventario"):
        return redirect('login')

    depositos = list(Deposito.objects.order_by('nombre').values_list('nombre', flat=True))
    deposito_query = (request.GET.get('deposito') or '').strip()

    data = [
        {
            'NumerodeCliente': v.NumerodeCliente,
            'marca': v.marca,
            'modelo': v.modelo,
            'anio': v.anio,
            'placas': v.placas,
            'vin': v.vin,
            'deposito': v.deposito,
            'motivo': v.motivo,
            'grua_motivo': v.grua_motivo,
            'grua_direccion': v.grua_direccion,
            'tipo': v.tipo_servicio,
            'estatus': v.estatus_legal,
            'fecha': v.fecha_ingreso.isoformat() if v.fecha_ingreso else '',
        }
        for v in Vehiculo.objects.all()
    ]
    return render(
        request,
        'Vehiculos/vehiculos.html',
        {
            'vehiculos_data': data,
            'can_registrar': _has_permission(request, "registrar"),
            'can_liberar': _has_permission(request, "liberar"),
            'can_solicitar_correccion': _has_permission(request, "solicitar_correccion"),
            'depositos': depositos,
            'deposito_actual': deposito_query,
        },
    )


CLIENTE_FIELD_LABELS = {
    'fecha_registro': 'Fecha de registro',
    'sap': 'Codigo SAP',
    'nombre': 'Nombre del lugar',
    'tipo_cuenta': 'Tipo de cuenta',
    'lista_precios': 'Lista de precios',
    'numero_empleado': 'Código 6 dígitos',
    'latitud': 'Latitud',
    'longitud': 'Longitud',
    'direccion': 'Direccion',
    'calle': 'Calle',
    'colonia': 'Colonia',
    'municipio': 'Municipio',
    'codigo_postal': 'Codigo postal',
    'zona': 'Zona',
    'estado': 'Estado',
    'poblacion': 'Poblacion',
}

from decimal import Decimal # 👈 Agregamos esto para manejar bien los números en base de datos

def operadorregistrador_view(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "operadorregistrador"):
        return _reject_unauthorized(request)

    rol = _get_role(request) 

    def build_context(values=None):
        return {
            'form_values': values or {},
            'rol': rol,
            'google_maps_api_key': os.getenv('GOOGLE_MAPS_API_KEY', '').strip(),
        }

    if request.method == 'POST':
        post = request.POST
        values = {key: (post.get(key) or '').strip() for key in CLIENTE_FIELD_LABELS}

        # 🚀 AQUÍ GENERAMOS EL SAP AUTOMÁTICAMENTE
        values['sap'] = _generar_codigo_sap()

        # Agregar campos de visita
        values['frecuencia_visita'] = post.get('frecuencia_visita', '').strip()
        values['dias_visita'] = ','.join(post.getlist('dias_visita'))

        # 🔒 BLOQUEO BACKEND PARA PROMOTOR
        if rol == ROLE_PROMOTOR:
            values['tipo_cuenta'] = Cliente.TIPO_PROSPECTO
            values['lista_precios'] = 'DEFAULT'
            values['zona'] = ''
            values['estado'] = ''
            values['poblacion'] = ''

        # ✅ CAMPOS OBLIGATORIOS DINÁMICOS (Quitamos 'sap' porque ya lo genera el sistema)
        if rol == ROLE_PROMOTOR:
            required = ['fecha_registro', 'nombre', 'numero_empleado', 'latitud', 'longitud']
        else:
            required = ['fecha_registro', 'nombre', 'numero_empleado', 'tipo_cuenta', 
                        'latitud', 'longitud', 'direccion', 'zona', 'estado', 'poblacion']

        missing = [field for field in required if not values.get(field)]
        if missing:
            labels = [CLIENTE_FIELD_LABELS[field] for field in missing]
            messages.error(request, f'Completa los campos obligatorios: {", ".join(labels)}.')
            return render(request, 'Vehiculos/operador.html', build_context(values))

        if not re.fullmatch(r"\d{6}", values['numero_empleado']):
            messages.error(request, 'El código de 6 dígitos debe contener exactamente 6 números.')
            return render(request, 'Vehiculos/operador.html', build_context(values))

        if Cliente.objects.filter(numero_empleado=values['numero_empleado']).exists():
            messages.error(request, 'El código de 6 dígitos ya está en uso. Elige uno diferente.')
            return render(request, 'Vehiculos/operador.html', build_context(values))

        # 🔒 VALIDACIÓN SOLO PARA NO PROMOTOR
        if rol != ROLE_PROMOTOR:
            if values['tipo_cuenta'] not in (Cliente.TIPO_DIRECTO, Cliente.TIPO_PROSPECTO):
                messages.error(request, 'Selecciona un tipo de cuenta valido.')
                return render(request, 'Vehiculos/operador.html', build_context(values))

        # Tomamos el SAP que generó el sistema
        sap = values['sap'].upper()

        fecha = parse_date(values['fecha_registro'])
        if fecha is None:
            messages.error(request, 'La fecha de registro no es valida.')
            return render(request, 'Vehiculos/operador.html', build_context(values))

        try:
            latitud = float(values['latitud'])
            longitud = float(values['longitud'])
        except ValueError:
            messages.error(request, 'Latitud y longitud deben ser numeros.')
            return render(request, 'Vehiculos/operador.html', build_context(values))

        usuario_actual = _get_current_user(request)
        operador_destino = usuario_actual
        if rol == ROLE_PROMOTOR:
            perfil = PerfilUsuario.objects.filter(user=usuario_actual).first()
            if not perfil or not perfil.operador_asignado:
                messages.error(request, 'Tu cuenta de promotor no tiene operador asignado.')
                return render(request, 'Vehiculos/operador.html', build_context(values))
            # Guardamos el cliente como creado por el promotor, no por el operador.
            operador_destino = usuario_actual

        cliente = Cliente.objects.create(
            sap=sap,
            numero_empleado=values['numero_empleado'],
            nombre=values['nombre'],
            tipo_cuenta=values['tipo_cuenta'],
            lista_precios=values['lista_precios'].upper(),
            latitud=latitud,
            longitud=longitud,
            direccion=values['direccion'],
            calle=values['calle'],
            colonia=values['colonia'],
            municipio=values['municipio'],
            codigo_postal=values['codigo_postal'],
            zona=values['zona'].upper(),
            estado=values['estado'].upper(),
            poblacion=values['poblacion'].upper(),
            fecha_registro=fecha,
            operador=operador_destino,
            frecuencia_visita=values['frecuencia_visita'],
            dias_visita=values['dias_visita'],
        )

        messages.success(request, f'Cliente {cliente.sap} - {cliente.nombre} registrado correctamente.')
        return redirect('clientes_list')

    return render(request, 'Vehiculos/operador.html', build_context())


def geolocalizacion_view(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "operadorregistrador") and not _has_permission(request, "gestionar_usuarios"):
        return _reject_unauthorized(request)

    sap_query = (request.GET.get('sap') or '').strip()
    clientes_qs = _scoped_clientes_queryset(request).select_related('operador')
    cliente = clientes_qs.filter(sap__iexact=sap_query).first() if sap_query else None

    if sap_query and not cliente:
        messages.error(request, f'No se encontró cliente con SAP "{sap_query}".')

    return render(
        request,
        'Vehiculos/geolocalizacion.html',
        {
            'sap_query': sap_query,
            'cliente': cliente,
            'rol': _get_role(request),
            'rol_label': ROLE_LABELS[_get_role(request)],
        },
    )



def clientes_list_view(request):
    if not _is_logged_in(request):
        return redirect('login')
    
    # 1. Usamos la función de seguridad que YA TIENES para filtrar por rol
    # Esto soluciona que el PROMOTOR vea todo.
    query = _scoped_clientes_queryset(request).select_related('operador').order_by('-fecha_registro', '-id')

    role = _get_role(request)
    user = _get_current_user(request)
    
    # 2. Permisos para el HTML
    can_borrar_clientes = _has_permission(request, "gestionar_usuarios")

    # --- 🚀 FILTROS DE BÚSQUEDA ---
    search = (request.GET.get('q') or '').strip()
    operador_id = request.GET.get('operador_filtro')
    tipo_filtro = request.GET.get('tipo_filtro')

    if search:
        query = query.filter(Q(sap__icontains=search) | Q(nombre__icontains=search))
    if operador_id:
        query = query.filter(operador_id=operador_id)
    if tipo_filtro:
        query = query.filter(tipo_cuenta=tipo_filtro)

    # --- 👤 LISTA DE USUARIOS PARA EL FILTRO (Solo para Jefes) ---
    operadores_lista = None
    if role in [ROLE_ADMIN, ROLE_ADMIN_MASTER]:
        User = get_user_model()
        operadores_lista = User.objects.filter(
            groups__name__in=[ROLE_OPERADOR, ROLE_ADMIN, ROLE_ADMIN_MASTER]
        ).distinct()

    # 3. Paginación
    paginator = Paginator(query, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
        'operador_id_actual': operador_id,
        'tipo_actual': tipo_filtro,
        'operadores_lista': operadores_lista,
        'rol': role,  # Enviará "administrador", "admin_master", etc.
        'rol_label': ROLE_LABELS.get(role, role),
        'can_borrar_clientes': can_borrar_clientes,
        'total_clientes': paginator.count,
    }
    return render(request, 'Vehiculos/clientes.html', context)


def importar_clientes_excel(request):
    if not _is_logged_in(request):
        return redirect("login")

    if not _has_permission(request, "operadorregistrador") and not _has_permission(request, "gestionar_usuarios"):
        return _reject_unauthorized(request)

    if request.method != "POST":
        return render(request, "Vehiculos/importar-clientes.html", {
            "required_columns": ["Nombre", "Tipo Cuenta", "Dirección", "Zona", "Estado", "Población", "Latitud", "Longitud", "Lista de Precios"],
            "optional_columns": ["Fecha de registro"],
        })

    archivo = request.FILES.get("excel")
    if not archivo:
        messages.error(request, "Selecciona un archivo .xlsx o .csv.")
        return redirect("importar_clientes_excel")

    # --- 1. LÓGICA PARA LEER EL ARCHIVO (Define 'rows') ---
    rows = []
    name = archivo.name.lower()
    
    if name.endswith(".csv"):
        import csv
        from io import TextIOWrapper
        encoding = 'utf-8-sig' # Soporta caracteres especiales de Excel
        try:
            text_stream = TextIOWrapper(archivo, encoding=encoding)
            reader = csv.reader(text_stream)
            rows = list(reader)
        except:
            messages.error(request, "Error al leer el CSV. Intenta guardarlo como UTF-8.")
            return redirect("importar_clientes_excel")
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            messages.error(request, f"Error al leer Excel: {e}")
            return redirect("importar_clientes_excel")

    if not rows:
        messages.error(request, "El archivo está vacío.")
        return redirect("importar_clientes_excel")

    # --- 2. PROCESAR CABECERAS (Define 'header_to_index') ---
    header_row = rows[0]
    raw_headers = [_normalize_excel_header(cell) for cell in header_row]
    canonical_headers = [_canonicalize_import_header(h) for h in raw_headers]
    header_to_index = {h: idx for idx, h in enumerate(canonical_headers) if h}

    # --- 3. PREPARAR SECUENCIA SAP ---
    last_sap = Cliente.objects.filter(sap__startswith="SAP-").order_by("-sap").values_list("sap", flat=True).first()
    last_num = 0
    if last_sap and "-" in last_sap:
        try:
            last_num = int(last_sap.split("-", 1)[1])
        except ValueError:
            last_num = 0
    
    next_seq_num = last_num + 1

    # --- 4. FUNCIÓN AUXILIAR (Define 'to_float') ---
    def to_float(value):
        if value is None or value == "":
            return 0.0
        try:
            # Limpia comas si vienen en formato europeo/latino
            text = str(value).strip().replace(",", ".")
            return float(text)
        except:
            return 0.0

    # --- 5. IMPORTACIÓN MASIVA ---
    created = 0
    errores = []
    current_user = _get_current_user(request)
    today = timezone.localdate()

    with transaction.atomic():
        for row_num, row in enumerate(rows[1:], start=2):
            if not row or all(cell in (None, "") for cell in row):
                continue

            def cell_value(col):
                idx = header_to_index.get(col)
                return row[idx] if idx is not None and idx < len(row) else None

            # Generamos el SAP de la secuencia
            sap_final = f"SAP-{next_seq_num:05d}"
            
            nombre = str(cell_value("nombre") or "").strip()
            if not nombre:
                continue # Saltamos filas sin nombre

            tipo_raw = str(cell_value("tipo_cuenta") or "").strip().upper()
            tipo = Cliente.TIPO_DIRECTO if tipo_raw.startswith("DIR") else Cliente.TIPO_PROSPECTO

            try:
                Cliente.objects.create(
                    sap=sap_final,
                    nombre=nombre,
                    tipo_cuenta=tipo,
                    lista_precios=str(cell_value("lista_precios") or "DEFAULT").strip().upper(),
                    latitud=to_float(cell_value("latitud")),
                    longitud=to_float(cell_value("longitud")),
                    direccion=str(cell_value("direccion") or "").strip(),
                    zona=str(cell_value("zona") or "").strip().upper(),
                    estado=str(cell_value("estado") or "").strip().upper(),
                    poblacion=str(cell_value("poblacion") or "").strip().upper(),
                    fecha_registro=_parse_excel_date(cell_value("fecha_registro")) or today,
                    operador=current_user,
                )
                created += 1
                next_seq_num += 1 # Solo sumamos si se creó con éxito
            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")

    if created:
        messages.success(request, f"¡Importación exitosa! Se crearon {created} clientes de SAP.")
    if errores:
        for err in errores[:5]: # Mostramos solo los primeros 5 errores para no saturar
            messages.error(request, err)

    return redirect("clientes_list")

CREDITO_FIELDS_NUMERICOS = [
    ('dias_maximos_entrega', 'Dias maximos de entrega', 'int'),
    ('holgura_dto_pp', 'Holgura Dto PP', 'decimal'),
    ('dias_para_fecha_entrega', 'Dias para fecha de entrega', 'int'),
]

CREDITO_FIELDS_BOOLEANOS = [
    ('pedido_excede_limite_credito', '¿Pedido puede exceder limite de credito?'),
    ('bloquear_cliente_factura_vencida', 'Bloquear cliente por factura vencida'),
    ('bloqueo_venta_documento_pendiente', 'Bloqueo venta por documento pendiente'),
    ('orden_compra_adquirida', 'Orden de compra adquirida'),
    ('permitir_devolucion', 'Permitir devolucion'),
    ('bloqueo_venta', 'Bloqueo de venta'),
    ('bloqueo_cheques_pendientes', 'Bloqueo de cheques pendientes'),
    ('tomar_inventario', '¿Tomar inventario?'),
    ('modificar_condicion_pago', 'Modificar la condicion de pago'),
    ('orden_compra_automatico', 'Orden de compra automatico (Auto Venta)'),
]


def editar_credito_view(request, cliente_id):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "gestionar_credito"):
        return _reject_unauthorized(request)
    if request.method != 'POST':
        return redirect('clientes_list')

    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        messages.error(request, 'Cliente no encontrado.')
        return redirect('clientes_list')

    errores = []

    for field, label, tipo in CREDITO_FIELDS_NUMERICOS:
        raw = (request.POST.get(field) or '').strip()
        if raw == '':
            setattr(cliente, field, 0)
            continue
        try:
            if tipo == 'int':
                value = int(raw)
                if value < 0:
                    raise ValueError
            else:
                value = float(raw)
                if value < 0:
                    raise ValueError
            setattr(cliente, field, value)
        except ValueError:
            errores.append(f'{label} debe ser un numero valido no negativo.')

    for field, _label in CREDITO_FIELDS_BOOLEANOS:
        setattr(cliente, field, request.POST.get(field) == 'on')

    if errores:
        for err in errores:
            messages.error(request, err)
        return redirect('clientes_list')

    cliente.save()
    messages.success(request, f'Panel de credito de {cliente.sap} actualizado correctamente.')
    return redirect('clientes_list')


MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre',
}


def historial_view(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "operadorregistrador"):
        return _reject_unauthorized(request)
    

    user = _get_current_user(request)
    if user is None:
        return redirect('login')

    clientes = (
        Cliente.objects.filter(operador=user)
        .order_by('-fecha_registro', '-id')
    )

    grupos = {}
    for cliente in clientes:
        key = (cliente.fecha_registro.year, cliente.fecha_registro.month)
        grupos.setdefault(key, []).append(cliente)

    meses = []
    for (anio, mes), items in sorted(grupos.items(), reverse=True):
        meses.append({
            'anio': anio,
            'mes': mes,
            'mes_nombre': MESES_ES[mes],
            'etiqueta': f'{MESES_ES[mes]} {anio}',
            'slug': f'{anio}-{mes:02d}',
            'total': len(items),
            'clientes': items,
        })

    context = {
        'meses': meses,
        'total_clientes': clientes.count(),
        'total_meses': len(meses),
        'rol': _get_role(request),
        'rol_label': ROLE_LABELS[_get_role(request)],
    }
    return render(request, 'Vehiculos/historial.html', context)


def vehiculos_list(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "ver_inventario"):
        return redirect('login')

    depositos = list(Deposito.objects.order_by('nombre').values_list('nombre', flat=True))
    deposito_query = (request.GET.get('deposito') or '').strip()

    data = [
        {
            'CodigoSap': v.CodigoSAp,
            'marca': v.marca,
            'modelo': v.modelo,
            'anio': v.anio,
            'placas': v.placas,
            'vin': v.vin,
            'deposito': v.deposito,
            'motivo': v.motivo,
            'grua_motivo': v.grua_motivo,
            'grua_direccion': v.grua_direccion,
            'tipo': v.tipo_servicio,
            'estatus': v.estatus_legal,
            'fecha': v.fecha_ingreso.isoformat() if v.fecha_ingreso else '',
        }
        for v in Vehiculo.objects.all()
    ]
    # return render(
    #     request,
    #     'Vehiculos/vehiculos.html',
    #     {
    #         'vehiculos_data': data,
    #         'can_registrar': _has_permission(request, "registrar"),
    #         'can_liberar': _has_permission(request, "liberar"),
    #         'can_solicitar_correccion': _has_permission(request, "solicitar_correccion"),
    #         'depositos': depositos,
    #         'deposito_actual': deposito_query,
    #     },
    # )


def liberar_vehiculo(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "liberar"):
        return _reject_unauthorized(request)

    def build_context(prefill=None):
        lookup_data = [
            {
                'folio': v.folio,
                'placas': v.placas,
                'vin': v.vin,
                'vehiculo': f'{v.marca} {v.modelo} {v.anio}',
            }
            for v in Vehiculo.objects.exclude(estatus_legal=Vehiculo.ESTATUS_LIBERADO)[:200]
        ]
        return {
            'vehiculos_lookup_data': lookup_data,
            'vehiculo_prefill': prefill,
            'can_liberar': _has_permission(request, "liberar"),
        }

    folio_query = (request.GET.get('folio') or '').strip().upper()
    vehiculo_prefill = None
    if folio_query:
        vehiculo_prefill = Vehiculo.objects.filter(folio=folio_query).first()

    if request.method == 'POST':
        post = request.POST
        folio = (post.get('folio') or '').strip().upper()
        placas = (post.get('placas') or '').strip().upper()
        vin = (post.get('vin') or '').strip().upper()
        oficio = (post.get('oficio') or '').strip()
        fecha_oficio = post.get('fecha_oficio')
        autoriza = (post.get('autoriza') or '').strip()
        observaciones = (post.get('observaciones') or '').strip()

        if not all([folio, placas, vin, oficio, fecha_oficio, autoriza]):
            messages.error(request, 'Completa todos los datos obligatorios de liberacion.')
            return render(request, 'Vehiculos/liberar-vehiculo.html', build_context(vehiculo_prefill))

        vehiculo = Vehiculo.objects.filter(folio=folio).first()
        if not vehiculo:
            messages.error(request, f'No existe un vehiculo con folio {folio}.')
            return render(request, 'Vehiculos/liberar-vehiculo.html', build_context(vehiculo_prefill))

        if vehiculo.placas != placas or vehiculo.vin != vin:
            messages.error(request, 'Placas o VIN no coinciden con el folio capturado.')
            return render(request, 'Vehiculos/liberar-vehiculo.html', build_context(vehiculo))

        vehiculo.oficio = oficio
        vehiculo.fecha_oficio = fecha_oficio
        vehiculo.autoriza_liberacion = autoriza
        vehiculo.liberacion_observaciones = observaciones
        vehiculo.documento_nombre = (
            request.FILES['documento_pdf'].name if request.FILES.get('documento_pdf') else vehiculo.documento_nombre
        )
        vehiculo.aceite_drenado = post.get('aceite_drenado') == 'on'
        vehiculo.combustible_drenado = post.get('combustible_drenado') == 'on'
        vehiculo.anticongelante_drenado = post.get('anticongelante_drenado') == 'on'
        vehiculo.sin_objetos_pendientes = post.get('sin_objetos_pendientes') == 'on'
        vehiculo.marcar_liberado()
        vehiculo.save()

        messages.success(request, f'Vehiculo {vehiculo.folio} liberado correctamente.')
        return redirect('vehiculos')

    return render(request, 'Vehiculos/liberar-vehiculo.html', build_context(vehiculo_prefill))


def logout_view(request):
    auth_logout(request)
    request.session.flush()
    return redirect('login')


def solicitar_correccion(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "solicitar_correccion"):
        return _reject_unauthorized(request)

    folio_query = (request.GET.get('folio') or '').strip().upper()
    vehiculo = Vehiculo.objects.filter(folio=folio_query).first() if folio_query else None

    if request.method == 'POST':
        folio = (request.POST.get('folio') or '').strip().upper()
        campo = (request.POST.get('campo') or '').strip()
        valor_nuevo = (request.POST.get('valor_nuevo') or '').strip()
        motivo = (request.POST.get('motivo') or '').strip()

        vehiculo = Vehiculo.objects.filter(folio=folio).first()
        if not vehiculo:
            messages.error(request, 'No se encontro el vehiculo solicitado.')
            return redirect('solicitar_correccion')

        if campo not in CORRECCION_FIELDS:
            messages.error(request, 'Selecciona un campo valido para corregir.')
            return redirect('solicitar_correccion')

        if not valor_nuevo:
            messages.error(request, 'Ingresa el valor correcto.')
            return redirect('solicitar_correccion')

        field = Vehiculo._meta.get_field(campo)
        if field.choices:
            valid_choices = [choice[0] for choice in field.choices]
            if valor_nuevo not in valid_choices:
                messages.error(request, 'El valor no es valido para el campo seleccionado.')
                return redirect('solicitar_correccion')

        try:
            _coerce_field_value(campo, valor_nuevo)
        except Exception:
            messages.error(request, 'El valor no tiene el formato correcto para ese campo.')
            return redirect('solicitar_correccion')

        SolicitudCorreccion.objects.create(
            vehiculo=vehiculo,
            solicitante=_get_current_user(request),
            campo=campo,
            valor_nuevo=valor_nuevo,
            motivo=motivo,
        )

        messages.success(request, 'Solicitud enviada al administrador.')
        return redirect('vehiculos')

    return render(
        request,
        'Vehiculos/solicitar-correccion.html',
        {
            'vehiculo': vehiculo,
            'campos': CORRECCION_FIELDS,
        },
    )


def solicitudes_correccion(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "gestionar_correcciones"):
        return _reject_unauthorized(request)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        solicitud_id = request.POST.get('solicitud_id')
        solicitud = SolicitudCorreccion.objects.select_related('vehiculo').filter(id=solicitud_id).first()

        if not solicitud:
            messages.error(request, 'No se encontro la solicitud.')
            return redirect('solicitudes_correccion')

        if solicitud.estatus != SolicitudCorreccion.ESTATUS_PENDIENTE:
            messages.error(request, 'La solicitud ya fue atendida.')
            return redirect('solicitudes_correccion')

        if action == 'rechazar':
            solicitud.estatus = SolicitudCorreccion.ESTATUS_RECHAZADA
            solicitud.resuelto_en = timezone.now()
            solicitud.resuelto_por = _get_current_user(request)
            solicitud.save(update_fields=['estatus', 'resuelto_en', 'resuelto_por'])
            messages.info(request, 'Solicitud rechazada.')
            return redirect('solicitudes_correccion')

        if action == 'aprobar':
            try:
                nuevo_valor = _coerce_field_value(solicitud.campo, solicitud.valor_nuevo)
            except Exception:
                messages.error(request, 'No se pudo aplicar el cambio por formato invalido.')
                return redirect('solicitudes_correccion')

            vehiculo = solicitud.vehiculo
            if solicitud.campo == 'estatus_legal':
                if nuevo_valor == Vehiculo.ESTATUS_LIBERADO:
                    vehiculo.marcar_liberado()
                else:
                    vehiculo.estatus_legal = nuevo_valor
                    vehiculo.liberado = False
                    vehiculo.fecha_liberacion = None
            else:
                setattr(vehiculo, solicitud.campo, nuevo_valor)
            vehiculo.save()

            solicitud.estatus = SolicitudCorreccion.ESTATUS_APROBADA
            solicitud.resuelto_en = timezone.now()
            solicitud.resuelto_por = _get_current_user(request)
            solicitud.save(update_fields=['estatus', 'resuelto_en', 'resuelto_por'])

            messages.success(request, 'Solicitud aplicada correctamente.')
            return redirect('solicitudes_correccion')

        messages.error(request, 'Accion invalida.')
        return redirect('solicitudes_correccion')

    solicitudes = list(
        SolicitudCorreccion.objects.select_related('vehiculo', 'solicitante').order_by('-creado_en')
    )
    for solicitud in solicitudes:
        solicitud.campo_label = CORRECCION_FIELDS.get(solicitud.campo, solicitud.campo)
    return render(
        request,
        'Vehiculos/solicitudes-correccion.html',
        {
            'solicitudes': solicitudes,
        },
    )


def solicitar_correccion_cliente(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "solicitar_correccion_cliente"):
        return _reject_unauthorized(request)

    sap_query = (request.GET.get('sap') or '').strip()
    clientes_qs = _scoped_clientes_queryset(request)
    cliente = clientes_qs.filter(sap__iexact=sap_query).first() if sap_query else None

    if request.method == 'POST':
        sap = (request.POST.get('sap') or '').strip()
        campo = (request.POST.get('campo') or '').strip()
        valor_nuevo = (request.POST.get('valor_nuevo') or '').strip()
        motivo = (request.POST.get('motivo') or '').strip()

        cliente = clientes_qs.filter(sap__iexact=sap).first()
        if not cliente:
            messages.error(request, 'No se encontro el cliente solicitado.')
            return redirect('solicitar_correccion_cliente')

        if campo not in CLIENTE_CORRECCION_FIELDS:
            messages.error(request, 'Selecciona un campo valido para corregir.')
            return redirect('solicitar_correccion_cliente')

        if not valor_nuevo:
            messages.error(request, 'Ingresa el valor correcto.')
            return redirect('solicitar_correccion_cliente')

        field = Cliente._meta.get_field(campo)
        if field.choices:
            valid_choices = [choice[0] for choice in field.choices]
            if valor_nuevo not in valid_choices:
                messages.error(request, 'El valor no es valido para el campo seleccionado.')
                return redirect('solicitar_correccion_cliente')

        try:
            _coerce_model_field_value(Cliente, campo, valor_nuevo)
        except Exception:
            messages.error(request, 'El valor no tiene el formato correcto para ese campo.')
            return redirect('solicitar_correccion_cliente')

        try:
            SolicitudCorreccionCliente.objects.create(
                cliente=cliente,
                solicitante=_get_current_user(request),
                campo=campo,
                valor_nuevo=valor_nuevo,
                motivo=motivo,
            )
        except ProgrammingError:
            messages.error(request, "Faltan migraciones de correcciones de clientes. Ejecuta: manage.py migrate")
            return redirect('clientes_list')
        messages.success(request, 'Solicitud enviada al administrador.')
        return redirect('clientes_list')

    return render(
        request,
        'Vehiculos/solicitar-correccion-cliente.html',
        {
            'cliente': cliente,
            'campos': CLIENTE_CORRECCION_FIELDS,
        },
    )


def solicitudes_correccion_clientes(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "gestionar_correcciones_clientes"):
        return _reject_unauthorized(request)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        solicitud_id = request.POST.get('solicitud_id')
        try:
            solicitud = (
                SolicitudCorreccionCliente.objects.select_related('cliente')
                .filter(id=solicitud_id)
                .first()
            )
        except ProgrammingError:
            messages.error(request, "Faltan migraciones de correcciones de clientes. Ejecuta: manage.py migrate")
            return redirect('dashboard')

        if not solicitud:
            messages.error(request, 'No se encontro la solicitud.')
            return redirect('solicitudes_correccion_clientes')

        if solicitud.estatus != SolicitudCorreccionCliente.ESTATUS_PENDIENTE:
            messages.error(request, 'La solicitud ya fue atendida.')
            return redirect('solicitudes_correccion_clientes')

        if action == 'rechazar':
            solicitud.estatus = SolicitudCorreccionCliente.ESTATUS_RECHAZADA
            solicitud.resuelto_en = timezone.now()
            solicitud.resuelto_por = _get_current_user(request)
            solicitud.save(update_fields=['estatus', 'resuelto_en', 'resuelto_por'])
            messages.info(request, 'Solicitud rechazada.')
            return redirect('solicitudes_correccion_clientes')

        if action == 'aprobar':
            try:
                nuevo_valor = _coerce_model_field_value(Cliente, solicitud.campo, solicitud.valor_nuevo)
            except Exception:
                messages.error(request, 'No se pudo aplicar el cambio por formato invalido.')
                return redirect('solicitudes_correccion_clientes')

            cliente = solicitud.cliente
            setattr(cliente, solicitud.campo, nuevo_valor)
            cliente.save()

            solicitud.estatus = SolicitudCorreccionCliente.ESTATUS_APROBADA
            solicitud.resuelto_en = timezone.now()
            solicitud.resuelto_por = _get_current_user(request)
            solicitud.save(update_fields=['estatus', 'resuelto_en', 'resuelto_por'])

            messages.success(request, 'Solicitud aplicada correctamente.')
            return redirect('solicitudes_correccion_clientes')

        messages.error(request, 'Accion invalida.')
        return redirect('solicitudes_correccion_clientes')

    try:
        solicitudes = list(
            SolicitudCorreccionCliente.objects.select_related('cliente', 'solicitante').order_by('-creado_en')
        )
    except ProgrammingError:
        messages.error(request, "Faltan migraciones de correcciones de clientes. Ejecuta: manage.py migrate")
        solicitudes = []
    for solicitud in solicitudes:
        solicitud.campo_label = CLIENTE_CORRECCION_FIELDS.get(solicitud.campo, solicitud.campo)

    return render(
        request,
        'Vehiculos/solicitudes-correccion-clientes.html',
        {
            'solicitudes': solicitudes,
        },
    )
def _normalize_excel_header(header):
    if header is None:
        return ""
    import unicodedata
    s = str(header).strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return s.replace(" ", "_").replace(".", "")

def borrar_masivo_clientes(request):
    if not _is_logged_in(request):
        return redirect('login')
    if not _has_permission(request, "gestionar_usuarios"):
        return _reject_unauthorized(request)

    if request.method == 'POST':
        ids_raw = request.POST.get('ids', '')
        if ids_raw:
            id_list = ids_raw.split(',')
            # Borrado eficiente en una sola consulta
            cantidad, _ = Cliente.objects.filter(id__in=id_list).delete()
            messages.success(request, f'Se eliminaron {cantidad} registros correctamente.')
        else:
            messages.warning(request, 'No se seleccionaron clientes para eliminar.')
            
    return redirect('clientes_list')



def exportar_clientes_csv(request):
    # 1. DETECCIÓN AGRESIVA DE ROL
    # Si eres superusuario (el que creaste en la terminal) o staff, ERES ADMIN.
    es_admin_maestro = request.user.is_superuser or request.user.is_staff
    
    # También revisamos la sesión por si acaso
    rol_sesion = request.session.get('rol', '').lower()
    if 'admin' in rol_sesion:
        es_admin_maestro = True

    print(f"--- DEBUG EXPORTACIÓN ---")
    print(f"Usuario: {request.user.username}")
    print(f"¿Es Admin detectado?: {es_admin_maestro}")

    # 2. FILTRADO DE DATOS
    if es_admin_maestro:
        # Si eres admin, no importa quién lo registró, te trae los 3, 100 o 1000 que existan
        queryset = Cliente.objects.all()
        print(f"Acción: Exportando TODO el inventario")
    else:
        # Si no eres admin, solo lo que tú picaste
        queryset = Cliente.objects.filter(operador=request.user)
        print(f"Acción: Exportando solo registros de {request.user.username}")

    print(f"Registros encontrados: {queryset.count()}")
    print(f"--------------------------")

    # 3. GENERACIÓN DEL ARCHIVO
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Respaldo_Clientes_Total.csv"'
    response.write(u'\ufeff'.encode('utf8'))

    writer = csv.writer(response)
    writer.writerow([
        'SAP ID', 'NOMBRE', 'TIPO CUENTA', 'LATITUD', 'LONGITUD', 
        'LISTA PRECIOS', 'CALLE', 'COLONIA', 'POBLACIÓN', 'MUNICIPIO', 
        'ESTADO', 'CP', 'ZONA', 'FECHA REGISTRO', 'REGISTRADO POR',
        'FRECUENCIA VISITA', 'DIAS VISITA'
    ])

    for c in queryset:
        # Buscamos el nombre del operador que lo registró originalmente
        responsable = "Sistema"
        if c.operador:
            # Intentamos traer su nombre real, si no, su email o username
            responsable = f"{c.operador.first_name} {c.operador.last_name}".strip() or c.operador.email or c.operador.username

        writer.writerow([
            c.sap,
            c.nombre,
            c.tipo_cuenta,
            c.latitud,
            c.longitud,
            c.lista_precios,
            getattr(c, 'calle', '-'),
            getattr(c, 'colonia', '-'),
            getattr(c, 'poblacion', '-'),
            getattr(c, 'municipio', '-'),
            getattr(c, 'estado', '-'),
            getattr(c, 'codigo_postal', '-'),
            getattr(c, 'zona', '-'),
            c.fecha_registro.strftime('%d/%m/%Y') if c.fecha_registro else '-',
            responsable,
            c.frecuencia_visita,
            c.dias_visita
        ])

    return response